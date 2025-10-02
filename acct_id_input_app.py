import logging
from logging.handlers import RotatingFileHandler
import re
from pathlib import Path
from email.message import EmailMessage

import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from openpyxl import load_workbook

# ========= External dependency you provide =========
# This must be implemented in your environment.
# It should return a pandas.DataFrame with columns in the same order as the SELECT.
from your_module import run_ccb_query  # <-- replace with the real import path


# =========================
# Config
# =========================
LOG_PATH = Path("account_app.log")
TEMPLATE_PATH = Path("account_template.xlsx")
OUTPUT_PATH = Path("account_report.xlsx")

APP_NAME = "Account Lookup & Export (DF/CCB)"


# =========================
# Logging
# =========================
logger = logging.getLogger("account_app")
logger.setLevel(logging.INFO)

# Rotate at ~1 MB, keep 5 backups
handler = RotatingFileHandler(LOG_PATH, maxBytes=1_000_000, backupCount=5, encoding="utf-8")
fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(name)s | %(message)s")
handler.setFormatter(fmt)
logger.addHandler(handler)

logger.info("=== Application start ===")


# =========================
# Utilities
# =========================
SAFE_ACCT_ID_PATTERN = re.compile(r"^[A-Za-z0-9_\-\.]{1,64}$")

def sanitize_acct_id(acct_id: str) -> str:
    """
    Enforce a conservative allowlist for acct_id to reduce injection risk since
    run_ccb_query expects a raw SQL string.
    """
    if not acct_id or not SAFE_ACCT_ID_PATTERN.match(acct_id):
        raise ValueError("acct_id contains invalid characters.")
    return acct_id

def build_select_query(acct_id: str) -> str:
    # NOTE: Because run_ccb_query takes a string, we avoid concatenating unsafe input.
    # We sanitize acct_id first (allowlist). If your CCB layer supports parameters,
    # switch to parameterized queries there.
    return f"""
        SELECT
            acct_id,
            name,
            segment,
            balance,
            status
        FROM accounts
        WHERE acct_id = '{acct_id}'
    """.strip()

def df_to_tuples(df: pd.DataFrame):
    """Return (columns, rows_as_tuples) in display-friendly types."""
    cols = list(df.columns)
    # Ensure pure Python types for Tk/Excel
    rows = [tuple(None if pd.isna(v) else (v.item() if hasattr(v, "item") else v) for v in row)
            for row in df.itertuples(index=False, name=None)]
    return cols, rows


# =========================
# Excel helpers (template-based)
# =========================
def ensure_template_with_headers(headers):
    """Create a very simple template if none exists."""
    if not TEMPLATE_PATH.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        for i, col in enumerate(headers, start=1):
            ws.cell(row=1, column=i, value=col)
        wb.save(TEMPLATE_PATH)
        logger.info(f"Template created at {TEMPLATE_PATH.resolve()} with headers: {headers}")

def write_df_to_template(df: pd.DataFrame, template_path: Path = TEMPLATE_PATH, output_path: Path = OUTPUT_PATH) -> Path:
    if df is None or df.empty:
        raise ValueError("No data to write to template.")
    headers = list(df.columns)
    ensure_template_with_headers(headers)

    wb = load_workbook(template_path)
    if "Report" not in wb.sheetnames:
        ws = wb.active
        ws.title = "Report"
    ws = wb["Report"]

    # header row: ensure matches df columns
    for i, col in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=col)

    # clear data rows
    if ws.max_row > 1:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

    # write data
    for r_idx, (_, series) in enumerate(df.iterrows(), start=2):
        for c_idx, col in enumerate(headers, start=1):
            val = series[col]
            ws.cell(row=r_idx, column=c_idx, value=None if pd.isna(val) else val)

    wb.save(output_path)
    logger.info(f"Excel written to {output_path.resolve()} (rows={len(df)}, cols={len(headers)})")
    return output_path


# =========================
# Email draft helper
# =========================
def create_email_draft(recipient, subject, body, attachment_path: Path, from_addr="you@example.com") -> Path:
    msg = EmailMessage()
    msg["From"] = from_addr
    msg["To"] = recipient
    msg["Subject"] = subject
    msg.set_content(body)

    with open(attachment_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=attachment_path.name,
        )

    draft_path = Path("draft_email.eml")
    with open(draft_path, "wb") as f:
        f.write(msg.as_bytes())

    logger.info(f"Email draft created at {draft_path.resolve()} (to={recipient}, subject={subject})")
    return draft_path


# =========================
# Tkinter UI
# =========================
class AccountLookupApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("980x680")

        # --- Inputs ---
        top = ttk.Frame(self, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="acct_id:").grid(row=0, column=0, sticky="w")
        self.acct_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.acct_var, width=30).grid(row=0, column=1, padx=8, pady=4, sticky="w")

        ttk.Button(top, text="Run Query", command=self.run_query).grid(row=0, column=2, padx=6)
        ttk.Button(top, text="Export → Excel + Draft Email", command=self.export_and_draft).grid(row=0, column=3, padx=6)

        # Email fields
        ttk.Label(top, text="To:").grid(row=1, column=0, sticky="w", pady=(6, 0))
        self.to_var = tk.StringVar(value="manager@example.com")
        ttk.Entry(top, textvariable=self.to_var, width=30).grid(row=1, column=1, padx=8, pady=(6, 0), sticky="w")

        ttk.Label(top, text="Subject:").grid(row=1, column=2, sticky="e", pady=(6, 0))
        self.subj_var = tk.StringVar(value="Account Report")
        ttk.Entry(top, textvariable=self.subj_var, width=40).grid(row=1, column=3, padx=6, pady=(6, 0), sticky="w")

        # Template paths
        path_frame = ttk.Frame(self, padding=(10, 0))
        path_frame.pack(fill="x")
        self.template_label_var = tk.StringVar(value=f"Template: {TEMPLATE_PATH.resolve()}")
        self.output_label_var = tk.StringVar(value=f"Output: {OUTPUT_PATH.resolve()}")
        ttk.Label(path_frame, textvariable=self.template_label_var).grid(row=0, column=0, sticky="w", pady=(6, 0))
        ttk.Label(path_frame, textvariable=self.output_label_var).grid(row=1, column=0, sticky="w", pady=(2, 8))
        ttk.Button(path_frame, text="Change Template...", command=self.pick_template).grid(row=0, column=1, padx=8, sticky="e")

        # --- SQL shown ---
        ttk.Label(self, text="Executed SQL:").pack(anchor="w", padx=10)
        self.query_text = tk.Text(self, height=4, wrap="word", bg="#f7f7f7")
        self.query_text.pack(fill="x", padx=10, pady=5)

        # --- Table (from DataFrame) ---
        ttk.Label(self, text="Query Result (table):").pack(anchor="w", padx=10)
        self.tree = ttk.Treeview(self, columns=(), show="headings", height=10)
        self.tree.pack(fill="x", padx=10, pady=5)
        vsb = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.place(in_=self.tree, relx=1.0, rely=0, relheight=1.0, anchor="ne")

        # --- Formatted details ---
        ttk.Label(self, text="Formatted Result:").pack(anchor="w", padx=10)
        self.result_text = tk.Text(self, height=10, wrap="word", bg="#eef6ff")
        self.result_text.pack(fill="both", expand=True, padx=10, pady=5)

        # --- Status ---
        self.status_var = tk.StringVar(value="Ready.")
        ttk.Label(self, textvariable=self.status_var, anchor="w").pack(fill="x", side="bottom")

        # data cache
        self.last_df: pd.DataFrame | None = None

        logger.info("UI initialized")

    # ----- UI helpers -----
    def set_columns(self, columns):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = columns if columns else ()
        for c in columns:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=160, anchor="w")

    def populate_rows(self, rows):
        self.tree.delete(*self.tree.get_children())
        for row in rows:
            self.tree.insert("", "end", values=row)

    def show_query(self, query):
        self.query_text.delete("1.0", "end")
        self.query_text.insert("1.0", query.strip())

    def show_formatted(self, cols, rows):
        self.result_text.delete("1.0", "end")
        if not rows:
            self.result_text.insert("1.0", "No results found.")
            return
        # Pretty key-value blocks
        for row in rows:
            block = "\n".join(f"{c}: {v}" for c, v in zip(cols, row))
            self.result_text.insert("end", block + "\n" + "-"*56 + "\n")

    # ----- Actions -----
    def run_query(self):
        acct_id_raw = self.acct_var.get().strip()
        if not acct_id_raw:
            messagebox.showinfo("Input required", "Please enter an acct_id.")
            return

        try:
            acct_id = sanitize_acct_id(acct_id_raw)
        except ValueError as e:
            logger.warning(f"acct_id validation failed: {acct_id_raw!r} | {e}")
            messagebox.showerror("Invalid acct_id", str(e))
            return

        query = build_select_query(acct_id)
        self.show_query(query)
        logger.info(f"Running query for acct_id={acct_id}")

        try:
            df = run_ccb_query(query)  # <-- returns pandas.DataFrame
        except Exception as e:
            logger.exception(f"run_ccb_query failed: {e}")
            messagebox.showerror("Query failed", f"Query failed:\n{e}")
            self.status_var.set("Error.")
            return

        if df is None or df.empty:
            logger.info("Query returned no rows.")
            self.last_df = None
            self.set_columns(["Message"])
            self.populate_rows([("No results found.",)])
            self.show_formatted([], [])
            self.status_var.set("No results.")
            return

        self.last_df = df
        cols, rows = df_to_tuples(df)
        logger.info(f"Query returned {len(rows)} row(s), columns={cols}")

        self.set_columns(cols)
        self.populate_rows(rows)
        self.show_formatted(cols, rows)
        self.status_var.set(f"Found {len(rows)} row(s).")

    def export_and_draft(self):
        if self.last_df is None or self.last_df.empty:
            # Try running the query with current acct_id
            self.run_query()
            if self.last_df is None or self.last_df.empty:
                return

        try:
            out_path = write_df_to_template(self.last_df, TEMPLATE_PATH, OUTPUT_PATH)
            self.output_label_var.set(f"Output: {Path(out_path).resolve()}")

            to = self.to_var.get().strip()
            subject = self.subj_var.get().strip() or "Account Report"
            acct_id = self.acct_var.get().strip() or "(unspecified)"
            body = f"""Hello,

Please find attached the report for account {acct_id}.

Regards,
Team
"""
            draft_path = create_email_draft(
                recipient=to,
                subject=subject,
                body=body,
                attachment_path=out_path,
                from_addr="you@example.com",
            )

            messagebox.showinfo(
                "Success",
                f"Excel saved to:\n{out_path}\n\nDraft email saved to:\n{draft_path}\n\n"
                "Open the .eml in your mail client to review/send."
            )
            self.status_var.set("Exported and draft created.")
        except Exception as e:
            logger.exception(f"Export or draft failed: {e}")
            messagebox.showerror("Error", f"Export or draft failed:\n{e}")
            self.status_var.set("Error during export.")

    def pick_template(self):
        path = filedialog.askopenfilename(
            title="Choose Excel Template",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if path:
            global TEMPLATE_PATH
            TEMPLATE_PATH = Path(path)
            self.template_label_var.set(f"Template: {TEMPLATE_PATH.resolve()}")
            logger.info(f"Template changed to {TEMPLATE_PATH.resolve()}")


# =========================
# Run app
# =========================
if __name__ == "__main__":
    try:
        app = AccountLookupApp()
        app.mainloop()
    except Exception as e:
        logger.exception(f"Fatal error in mainloop: {e}")
        raise
    finally:
        logger.info("=== Application exit ===")
